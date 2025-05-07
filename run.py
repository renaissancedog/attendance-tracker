# Code based off of and data pulled from Pi Day event
# Note that all attendance values in the pointlogger for the event are marked as False.
# To reuse, update eventval and eventtype
from openpyxl import load_workbook

eventval=4
eventtype="Event"
responses = load_workbook("formresponses.xlsx", data_only=True)
responsesheet=responses["Form Responses 1"]
loggerbook=load_workbook("pointlogger.xlsx")
logger=loggerbook[eventtype]
res_length = sum(1 for row in responsesheet.iter_rows(min_col=1, max_col=1, values_only=True) if row[0] is not None)
log_length = sum(1 for row in logger.iter_rows(min_col=1, max_col=1, values_only=True) if row[0] is not None)
nameset=set()
loggednames=set()

for i in range(4, log_length+2):
    logger.cell(i,eventval).value=False
for i in range(2,res_length+1):
    name=(responsesheet.cell(i,4).value).strip()+" "+(responsesheet.cell(i,3).value).strip()
    nameset.add(name)
    for j in range(4, log_length+2):
        if (name.lower().strip()==logger.cell(j,1).value.lower().strip()):
            logger.cell(j,eventval).value=True
            nameset.remove(name)
            loggednames.add(name)
print(loggednames, len(loggednames), "logged")
print(nameset,len(nameset), "missing, log by hand")
loggerbook.save(filename="pointlogger.xlsx")